import csv
from contextlib import contextmanager
from pathlib import Path
from typing import Generator, Iterable

import pdfkit
from inscriptis import get_text
from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill
from openpyxl.worksheet import Worksheet
from pyramid.request import Request
from pyramid_jinja2 import get_jinja2_environment
from xvfbwrapper import Xvfb

from .models import Amendement, Lecture

PDFKIT_OPTIONS = {"quiet": ""}
STATIC_PATH = Path(__file__).parent / "static"
PDF_CSS = str(STATIC_PATH / "css" / "print.css")

FIELDS_NAMES = {
    "article": "Num article",
    "article_titre": "Titre article",
    "article_order": "Ordre article",
    "alinea": "Alinéa",
    "num": "Num amdt",
    "auteur": "Auteur(s)",
    "date_depot": "Date de dépôt",
    "id_discussion_commune": "Discussion commune ?",
    "identique": "Identique ?",
    "parent": "Parent",
    "dispositif": "Corps amdt",
    "objet": "Exposé amdt",
    "observations": "Objet amdt",
    "reponse": "Réponse",
    "comments": "Commentaires",
    "avis": "Avis du Gouvernement",
}


def rename_field(field_name: str) -> str:
    return FIELDS_NAMES.get(field_name, field_name.capitalize())


EXCLUDED_FIELDS = [
    "Amendement",
    "pk",
    "article_pk",
    "lecture_pk",
    "parent_pk",
    "parent_rectif",
    "subdiv_type",
    "subdiv_titre",
    "subdiv_num",
    "subdiv_mult",
    "subdiv_pos",
    "created_at",
    "modified_at",
    "bookmarked_at",
]
FIELDS = [
    field
    for field in Amendement.__table__.columns.keys()
    if field not in EXCLUDED_FIELDS
] + [
    "article",
    "article_titre",
    "article_order",
    "parent",
    "gouvernemental",
    "chambre",
    "num_texte",
    "organe",
    "session",
]


HEADERS = [rename_field(field_name) for field_name in FIELDS]


DARK_BLUE = Color(rgb="00182848")
WHITE = Color(rgb="00FFFFFF")


def write_csv(lecture: Lecture, filename: str, request: Request) -> int:
    nb_rows = 0
    with open(filename, "w", encoding="utf-8-sig") as file_:
        file_.write(";".join(HEADERS) + "\n")
        writer = csv.DictWriter(
            file_,
            fieldnames=FIELDS,
            delimiter=";",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        for amendement in sorted(lecture.amendements):
            writer.writerow(export_amendement(amendement))
            nb_rows += 1
    return nb_rows


@contextmanager
def xvfb_if_supported() -> Generator:
    try:
        with Xvfb():
            yield
    except (EnvironmentError, OSError, RuntimeError):
        yield


def generate_html_for_pdf(request: Request, template_name: str, context: dict) -> str:
    """Mostly useful for testing purpose."""
    env = get_jinja2_environment(request, name=".html")
    template = env.get_template(template_name)
    content: str = template.render(**context)
    return content


def write_pdf(lecture: Lecture, filename: str, request: Request) -> None:
    content = generate_html_for_pdf(request, "print.html", {"lecture": lecture})
    with xvfb_if_supported():
        pdfkit.from_string(content, filename, options=PDFKIT_OPTIONS, css=PDF_CSS)


def write_pdf1(
    lecture: Lecture, amendement: Amendement, filename: str, request: Request
) -> None:
    content = generate_html_for_pdf(
        request,
        "print1.html",
        {"amendement": amendement, "similaires": amendement.similaires},
    )
    with xvfb_if_supported():
        pdfkit.from_string(content, filename, options=PDFKIT_OPTIONS, css=PDF_CSS)


def write_xlsx(lecture: Lecture, filename: str, request: Request) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amendements"

    _write_header_row(ws)
    nb_rows = _write_data_rows(ws, sorted(lecture.amendements))
    wb.save(filename)
    return nb_rows


def _write_header_row(ws: Worksheet) -> None:
    for column, value in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=column)
        cell.value = value
        cell.fill = PatternFill(patternType="solid", fgColor=DARK_BLUE)
        cell.font = Font(color=WHITE, sz=8)


def _write_data_rows(ws: Worksheet, amendements: Iterable[Amendement]) -> int:
    nb_rows = 0
    for amend in amendements:
        amend_dict = {rename_field(k): v for k, v in export_amendement(amend).items()}
        for column, value in enumerate(HEADERS, 1):
            cell = ws.cell(row=nb_rows + 2, column=column)
            cell.value = amend_dict[value]
            cell.font = Font(sz=8)
        nb_rows += 1
    return nb_rows


HTML_FIELDS = ["objet", "dispositif", "observations", "reponse", "comments"]


def export_amendement(amendement: Amendement) -> dict:
    data: dict = amendement.asdict(full=True)
    for field_name in HTML_FIELDS:
        if data[field_name] is not None:
            data[field_name] = html_to_text(data[field_name])
    for excluded_field in EXCLUDED_FIELDS:
        if excluded_field in data.keys():
            del data[excluded_field]
    return data


def html_to_text(html: str) -> str:
    text: str = get_text(html).strip()
    return text
