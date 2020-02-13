from collections import Counter
from typing import Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyramid.request import Request

from zam_repondeur.models import Amendement, Lecture

from .spreadsheet import FIELDS, HEADERS, export_amendement_for_spreadsheet


def write_xlsx(
    lecture: Lecture,
    filename: str,
    request: Request,
    amendements: Optional[List[Amendement]] = None,
) -> Counter:
    return _write_xlsx_amendements(
        filename, amendements or sorted(lecture.amendements),
    )


def _write_xlsx_amendements(
    filename: str, amendements: Iterable[Amendement],
) -> Counter:
    amendement_dicts = (
        {FIELDS[k]: v for k, v in export_amendement_for_spreadsheet(amendement).items()}
        for amendement in amendements
    )
    return _write_xlsx_amendement_dicts(filename, amendement_dicts)


def _write_xlsx_amendement_dicts(
    filename: str, amendement_dicts: Iterable[dict],
) -> Counter:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amendements"
    _write_xslsx_header_row(ws)
    counter = _export_xlsx_data_rows(ws, amendement_dicts)
    wb.save(filename)
    return counter


def _write_xslsx_header_row(ws: Worksheet) -> None:
    for column, value in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=column)
        cell.value = value


def _export_xlsx_data_rows(ws: Worksheet, amendement_dicts: Iterable[dict]) -> Counter:
    counter = Counter({"amendements": 0})
    for amend_dict in amendement_dicts:
        for column, value in enumerate(HEADERS, 1):
            cell = ws.cell(row=counter["amendements"] + 2, column=column)
            cell.value = amend_dict[value]
        counter["amendements"] += 1
    return counter
