import csv
from typing import Any, Iterable, Optional

from inscriptis import get_text
from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill
from openpyxl.worksheet import Worksheet
from pyramid.request import Request

from zam_repondeur.models import Amendement, Lecture

from .common import EXCLUDED_FIELDS


FIELDS = [
    field
    for field in Amendement.__table__.columns.keys()
    if field not in EXCLUDED_FIELDS
] + [
    "avis",
    "objet",
    "reponse",
    "comments",
    "article",
    "article_titre",
    "article_order",
    "parent",
    "gouvernemental",
    "chambre",
    "num_texte",
    "organe",
    "session",
    "affectation_email",
    "affectation_name",
]


FIELD_TO_COLUMN_NAME = {
    "article": "Num article",
    "article_titre": "Titre article",
    "article_order": "Ordre article",
    "alinea": "Alinéa",
    "num": "Num amdt",
    "auteur": "Auteur",
    "date_depot": "Date de dépôt",
    "id_discussion_commune": "Identifiant discussion commune",
    "id_identique": "Identifiant identique",
    "parent": "Parent (sous-amdt)",
    "corps": "Corps amdt",
    "expose": "Exposé amdt",
    "objet": "Objet amdt",
    "reponse": "Réponse",
    "comments": "Commentaires",
    "avis": "Avis du Gouvernement",
    "affectation_email": "Affectation (email)",
    "affectation_name": "Affectation (nom)",
}


def field_to_column_name(field_name: str) -> str:
    return FIELD_TO_COLUMN_NAME.get(field_name, field_name.capitalize())


COLUMN_NAME_TO_FIELD = {col: attr for attr, col in FIELD_TO_COLUMN_NAME.items()}


def column_name_to_field(column_name: str) -> Optional[str]:
    return COLUMN_NAME_TO_FIELD.get(column_name)


HEADERS = [field_to_column_name(field_name) for field_name in FIELDS]


HTML_FIELDS = ["corps", "expose", "objet", "reponse", "comments"]


DARK_BLUE = Color(rgb="00182848")
WHITE = Color(rgb="00FFFFFF")


def write_csv(lecture: Lecture, filename: str, request: Request) -> int:
    nb_rows = 0
    with open(filename, "w", encoding="utf-8-sig") as file_:
        file_.write(";".join(HEADERS) + "\n")
        writer = csv.DictWriter(
            file_,
            fieldnames=FIELDS,
            delimiter=";",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        for amendement in sorted(lecture.amendements):
            writer.writerow(export_amendement_for_spreadsheet(amendement))
            nb_rows += 1
    return nb_rows


def write_xlsx(lecture: Lecture, filename: str, request: Request) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amendements"

    _write_xslsx_header_row(ws)
    nb_rows = _write_xlsx_data_rows(ws, sorted(lecture.amendements))
    wb.save(filename)
    return nb_rows


def _write_xslsx_header_row(ws: Worksheet) -> None:
    for column, value in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=column)
        cell.value = value
        cell.fill = PatternFill(patternType="solid", fgColor=DARK_BLUE)
        cell.font = Font(color=WHITE, sz=8)


def _write_xlsx_data_rows(ws: Worksheet, amendements: Iterable[Amendement]) -> int:
    nb_rows = 0
    for amend in amendements:
        amend_dict = {
            field_to_column_name(k): v
            for k, v in export_amendement_for_spreadsheet(amend).items()
        }
        for column, value in enumerate(HEADERS, 1):
            cell = ws.cell(row=nb_rows + 2, column=column)
            cell.value = amend_dict[value]
            cell.font = Font(sz=8)
        nb_rows += 1
    return nb_rows


def export_amendement_for_spreadsheet(amendement: Amendement) -> dict:
    data: dict = amendement.asdict()
    for field_name in HTML_FIELDS:
        if data[field_name] is not None:
            data[field_name] = html_to_text(data[field_name])
    for excluded_field in EXCLUDED_FIELDS:
        if excluded_field in data.keys():
            del data[excluded_field]
    return {k: convert_boolean(v) for k, v in data.items()}


def convert_boolean(value: Any) -> Any:
    if value is True:
        return "Oui"
    elif value is False:
        return "Non"
    else:
        return value


def html_to_text(html: str) -> str:
    text: str = get_text(html).strip()
    return text
