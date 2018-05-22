import csv
from dataclasses import fields
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import (
    Color,
    Font,
    PatternFill,
)
from openpyxl.worksheet import Worksheet

from .models import Amendement


FIELDS = [field.name for field in fields(Amendement)]

FIELDS_NAMES = {
    'article': 'Nº article',
    'alinea': 'Alinéa',
    'num': 'Nº amdt ou sous-amdt',
    'auteur': 'Auteur(s)',
    'date_depot': 'Date de dépôt',
    'discussion_commune': 'Discussion commune ?',
    'identique': 'Identique ?',
}


DARK_BLUE = Color(rgb='00182848')
WHITE = Color(rgb='00FFFFFF')


def filter_dict(d, key_mapping):
    return {
        key_mapping[key]: value
        for key, value in d.items()
        if key in key_mapping
    }


def write_csv(amendements: Iterable[Amendement], filename: str) -> int:
    nb_rows = 0
    with open(filename, 'w', encoding='utf-8') as file_:
        writer = csv.DictWriter(
            file_,
            fieldnames=FIELDS,
            delimiter=';',
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()
        for amendement in amendements:
            writer.writerow(amendement.as_dict())
            nb_rows += 1
    return nb_rows


def write_xlsx(amendements: Iterable[Amendement], filename: str) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amendements"

    _write_header_row(ws)
    nb_rows = _write_data_row(ws, amendements)
    wb.save(filename)
    return nb_rows


def _write_header_row(ws: Worksheet) -> None:
    header_row = [
        FIELDS_NAMES.get(field, field.capitalize())
        for field in FIELDS
    ]
    for column, value in enumerate(header_row, 1):
        cell = ws.cell(row=1, column=column)
        cell.value = value
        cell.fill = PatternFill(
            patternType='solid',
            fgColor=DARK_BLUE,
        )
        cell.font = Font(
            color=WHITE,
            sz=8,
        )


def _write_data_row(ws, amendements) -> int:
    nb_rows = 0
    for amend in amendements:
        values = tuple(amend.as_dict().values())
        for column, value in enumerate(values, 1):
            cell = ws.cell(row=nb_rows + 2, column=column)
            cell.value = value
            cell.font = Font(sz=8)
        nb_rows += 1
    return nb_rows
