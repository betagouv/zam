import csv
from collections import Counter
from typing import Any, Iterable, Optional

from inscriptis import get_text
from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from pyramid.request import Request

from zam_repondeur.models import Amendement, Lecture

# NB: dict key order is used for spreadsheet columns order (Python 3.6+)
FIELDS = {
    "article": "Num article",
    "article_titre": "Titre article",
    "num": "Num amdt",
    "rectif": "Rectif",
    "parent": "Parent (sous-amdt)",
    "auteur": "Auteur",
    "groupe": "Groupe",
    "gouvernemental": "Gouvernemental",
    "corps": "Corps amdt",
    "expose": "Exposé amdt",
    "first_identique_num": "Identique",
    "avis": "Avis du Gouvernement",
    "objet": "Objet amdt",
    "reponse": "Réponse",
    "comments": "Commentaires",
    "affectation_email": "Affectation (email)",
    "affectation_name": "Affectation (nom)",
    "sort": "Sort",
}


COLUMN_NAME_TO_FIELD = {col: attr for attr, col in FIELDS.items()}


def column_name_to_field(column_name: str) -> Optional[str]:
    return COLUMN_NAME_TO_FIELD.get(column_name)


HEADERS = FIELDS.values()


HTML_FIELDS = ["corps", "expose", "objet", "reponse", "comments"]


def write_csv(lecture: Lecture, filename: str, request: Request) -> Counter:
    counter = Counter({"amendements": 0})
    with open(filename, "w", encoding="utf-8-sig") as file_:
        file_.write(";".join(HEADERS) + "\n")
        writer = csv.DictWriter(
            file_,
            fieldnames=list(FIELDS.keys()),
            delimiter=";",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        for amendement in sorted(lecture.amendements):
            writer.writerow(export_amendement_for_spreadsheet(amendement))
            counter["amendements"] += 1
    return counter


DARK_BLUE = Color(rgb="00182848")
WHITE = Color(rgb="00FFFFFF")


def write_xlsx(lecture: Lecture, filename: str, request: Request) -> Counter:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amendements"

    _write_xslsx_header_row(ws)
    counter = _write_xlsx_data_rows(ws, sorted(lecture.amendements))
    wb.save(filename)
    return counter


def _write_xslsx_header_row(ws: Worksheet) -> None:
    for column, value in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=column)
        cell.value = value
        cell.fill = PatternFill(patternType="solid", fgColor=DARK_BLUE)
        cell.font = Font(color=WHITE, sz=8)


def _write_xlsx_data_rows(ws: Worksheet, amendements: Iterable[Amendement]) -> Counter:
    counter = Counter({"amendements": 0})
    for amend in amendements:
        amend_dict = {
            FIELDS[k]: v for k, v in export_amendement_for_spreadsheet(amend).items()
        }
        for column, value in enumerate(HEADERS, 1):
            cell = ws.cell(row=counter["amendements"] + 2, column=column)
            cell.value = amend_dict[value]
            cell.font = Font(sz=8)
        counter["amendements"] += 1
    return counter


def export_amendement_for_spreadsheet(amendement: Amendement) -> dict:
    data: dict = {k: v for k, v in amendement.asdict().items() if k in FIELDS}
    for field_name in HTML_FIELDS:
        if data[field_name] is not None:
            data[field_name] = html_to_text(data[field_name])
    return {k: convert_boolean(v) for k, v in data.items()}


def convert_boolean(value: Any) -> Any:
    if value is True:
        return "Oui"
    elif value is False:
        return "Non"
    else:
        return value


def html_to_text(html: str) -> str:
    text: str = get_text(html).strip()
    return text
